import os, sys
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def main (file_name, step, column, parent_path):
    step_number = "step " + str(step)
    if step_number == "step 1":
        step_1(file_name, column, parent_path)
    elif step_number == "step 2":
        step_2(file_name, column, parent_path)
    elif step_number == "step 3":
        step_3(file_name, column, parent_path)
        step_3_KCM(file_name, column, parent_path)
        step_3_ecoli(file_name, column, parent_path)
        step_3_ligation(file_name, column, parent_path)
    elif step_number == "step 4":
        step_4(file_name, column, parent_path)
    elif step_number == "step 5":
        step_5(file_name, column, parent_path)
    elif step_number == "step 6":
        step_6(file_name, column, parent_path)
    elif step_number == "step 7":
        step_7(file_name, column, parent_path)
    elif step_number == "step 8":
        step_8(file_name, column, parent_path)
    elif step_number == "step 9":
        step_9(file_name, column, parent_path)
    elif step_number == "step 10":
        step_10(file_name, column, parent_path)
    elif step_number == "step 11":
        step_11(file_name, column, parent_path)
        step_11_KCM(file_name, column, parent_path)
        step_11_ecoli(file_name, column, parent_path)
    else:
        print("error")

def step_1(file_name, column, parent_path):
    ###workbook###
    workbook_path = parent_path
    workbook_path = str(workbook_path) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    ###volume###
    #Elution buffer
    elution_per_tube = float(ws["E92"].value)
    ws['E92'] = elution_per_tube / column
    elution_total = float(ws["E93"].value)
    ws['E93'] = elution_total / column
    #Buffer PM
    buffer_pm = float(ws["E105"].value)
    ws['E105'] = buffer_pm / column
    #Buffer PE
    buffer_pe_per_tube = float(ws["E107"].value)
    ws['E107'] = buffer_pe_per_tube / column
    buffer_pe_total = float(ws["E108"].value)
    ws['E108'] = buffer_pe_total / column
    #GXL PCR pre-mix
    template_dna = float(ws["E111"].value)
    ws['E111'] = template_dna / column
    reverse_primer =  float(ws["E112"].value)
    ws['E112'] = reverse_primer / column
    dNTP =  float(ws["E113"].value)
    ws['E113'] = dNTP / column
    fiveX_GXL_buffer =  float(ws["E114"].value)
    ws['E114'] = fiveX_GXL_buffer / column
    GXL_polymerase =  float(ws["E115"].value)
    ws['E115'] = GXL_polymerase / column
    DW =  float(ws["E116"].value)
    ws['E116'] = DW / column
    GXL_total =  float(ws["E117"].value)
    ws['E117'] = GXL_total / column
    ###tube###
        #no tube rules
    ###save excel###
    workbook.save(workbook_path)

def step_2(file_name, column, parent_path):
    ###workbook###
    workbook_path = parent_path
    workbook_path = str(workbook_path) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    ###volume###
    #elution buffer
    elution_buffer_per_tube =  float(ws["E92"].value)
    ws['E92'] = elution_buffer_per_tube / column
    elution_buffer_total =  float(ws["E93"].value)
    ws['E93'] = elution_buffer_total / column
    #enzyme cut
    tenX_green_fd_buffer =  float(ws["E96"].value)
    ws['E96'] = tenX_green_fd_buffer / column
    spel =  float(ws["E97"].value)
    ws['E97'] = spel / column
    DW =  float(ws["E98"].value)
    ws['E98'] = DW / column
    enzyme_per_tube =  float(ws["E99"].value)
    ws['E99'] = enzyme_per_tube / column
    enzyme_total =  float(ws["E100"].value)
    ws['E100'] = enzyme_total / column
    #buffer PM
    buffer_pm =  float(ws["E105"].value)
    ws['E105'] = buffer_pm / column
    #buffer PE
    buffer_pe_per_trough =  float(ws["E107"].value)
    ws['E107'] = buffer_pe_per_trough / column
    buffer_pe_total =  float(ws["E108"].value)
    ws['E108'] = buffer_pe_total / column
    ###tube###
    if 1 <= column <= 4:
        ws['F24'] = " "
        ws['F24'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
        ws['F23'] = " "
        ws['F23'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
        ws['F22'] = " "
        ws['F22'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    elif 5 <= column <= 8:
        ws['F24'] = " "
        ws['F24'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
        ws['F23'] = " "
        ws['F23'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    ###save excel###
    workbook.save(workbook_path)

def step_3_KCM(file_name, column, parent_path):
    ###workbook###
    workbook_path = parent_path
    workbook_path = str(workbook_path) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    if 1 <= column <= 6:
        ws['B28'] = " "
        ws['B28'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    ###save excel###
    workbook.save(workbook_path)
        
def step_3_ecoli(file_name, column, parent_path):
    ###workbook###
    workbook_path = parent_path
    workbook_path = str(workbook_path) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    if 1 <= column <= 6:
        ws['B30'] = " "
        ws['B30'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    ###save excel###
    workbook.save(workbook_path)
        
def step_3_ligation(file_name, column, parent_path):
    ###workbook###
    workbook_path = parent_path
    workbook_path = str(workbook_path) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    if 1 <= column <= 4:
        ws['F28'] = " "
        ws['F28'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
        ws['F27'] = " "
        ws['F27'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
        ws['F26'] = " "
        ws['F26'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    elif 5 <= column <= 8:
        ws['F28'] = " "
        ws['F28'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
        ws['F27'] = " "
        ws['F27'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    ###save excel###
    workbook.save(workbook_path)

def step_3(file_name, column, parent_path):
    ###workbook###
    workbook_path = parent_path
    workbook_path = str(workbook_path) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    ###volume###
    #DNA#
    #5X KCM
    fiveX_KCM_per_tube =  float(ws["E92"].value)
    ws['E92'] = fiveX_KCM_per_tube / column
    fiveX_KCM_total =  float(ws["E93"].value)
    ws['E93'] = fiveX_KCM_total / column
    #E.coli competent cell
    ecoli_per_tube =  float(ws["E95"].value)
    ws['E95'] = ecoli_per_tube / column
    ecoli_total =  float(ws["E96"].value)
    ws['E96'] = ecoli_total / column
    #ligation
    tenX_ligation =  float(ws["E99"].value)
    ws['E99'] = tenX_ligation / column
    T4_ligase =  float(ws["E100"].value)
    ws['E100'] = T4_ligase / column
    DW_tube =  float(ws["E101"].value)
    ws['E101'] = DW_tube / column
    ligation_pre_mix_per_tube =  float(ws["E102"].value)
    ws['E102'] = ligation_pre_mix_per_tube / column
    ligation_pre_mix_total =  float(ws["E103"].value)
    ws['E103'] = ligation_pre_mix_total / column
    #d.w.
    DW_trough =  float(ws["E110"].value)
    ws['E110'] = DW_trough / column
    #CELL#
    #LB medium
    LB_medium =  float(ws["E187"].value)
    ws['E187'] = LB_medium / column
    ###save excel###
    workbook.save(workbook_path)

def step_4(file_name, column, parent_path):
    ###workbook###
    workbook_path = parent_path
    workbook_path = str(workbook_path) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    ###volume###
    #dw
    dw = float(ws["E58"].value)
    ws['E58'] = dw / column
    #lb_medium
    lb_medium = float(ws["E61"].value)
    ws['E61'] = lb_medium / column
    ###tube###
        #no tube rules
    ###save excel###
    workbook.save(workbook_path)

def step_5(file_name, column, parent_path):
    ###workbook###
    workbook_path = parent_path
    workbook_path = str(workbook_path) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    ###volume###
    #DW
    dw_per_tube = float(ws["E91"].value)
    ws['E91'] = dw_per_tube / column
    dw_total =  float(ws["E92"].value)
    ws['E92'] = dw_total / column
    #buffer p1
    bufferp1 =  float(ws["E104"].value)
    ws['E104'] = bufferp1 / column
    #buffer p2
    bufferp2 =  float(ws["E106"].value)
    ws['E106'] = bufferp2 / column
    #buffer N3
    buffern3 =  float(ws["E108"].value)
    ws['E108'] = buffern3 / column
    #buffer pb
    bufferpb =  float(ws["E110"].value)
    ws['E110'] = bufferpb / column
    #buffer pe
    bufferpe_per_trough =  float(ws["E112"].value)
    ws['E112'] = bufferpe_per_trough / column
    bufferpe_total =  float(ws["E113"].value)
    ws['E113'] = bufferpe_total / column
    ###tube###
        #no tube rules
    ###save excel###
    workbook.save(workbook_path)

def step_6(file_name, column, parent_path):
    ###workbook###
    workbook_path = parent_path
    workbook_path = str(workbook_path) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    ###volume###
    #DNA#
        #no volume change
    #cell#
    BHISG = float(ws["E186"].value)
    ws['E186'] = BHISG / column
    ###tube###
        #no tube rules
    ###save excel###
    workbook.save(workbook_path)

def step_7(file_name, column, parent_path):
    ###workbook###
    workbook_path = parent_path
    workbook_path = str(workbook_path) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    ###volume###
    #BHIS
    BHIS_medium = float(ws["E64"].value)
    ws['E64'] = BHIS_medium / column
    #dw
    dw = float(ws["E66"].value)
    ws['E66'] = dw / column
    ###tube###
        #no tube rules
    ###save excel###
    workbook.save(workbook_path)

def step_8(file_name, column, parent_path):
    ###workbook###
    workbook_path = parent_path
    workbook_path = str(workbook_path) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    ###volume###
    #2X sapphire
    twoX_sapphire_per_tube = float(ws["E91"].value)
    ws['E91'] = twoX_sapphire_per_tube / column
    twoX_sapphire_total = float(ws["E92"].value)
    ws['E92'] = twoX_sapphire_total / column
    #buffer EB
    buffer_eb_per_tube = float(ws["E94"].value)
    ws['E94'] = buffer_eb_per_tube / column
    buffer_eb_total = float(ws["E95"].value)
    ws['E95'] = buffer_eb_total / column
    #buffer PM
    buffer_pm = float(ws["E104"].value)
    ws['E104'] = buffer_pm / column
    #buffer PE
    buffer_pe_per_tube = float(ws["E106"].value)
    ws['E106'] = buffer_pe_per_tube / column
    buffer_pe_total = float(ws["E107"].value)
    ws['E107'] = buffer_pe_total / column
    ###tube###
    if 1 <= column <= 6:
        ws['B22'] = " "
        ws['B22'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    ###save excel###
    workbook.save(workbook_path)

def step_9(file_name, column, parent_path):
    ###workbook###
    workbook_path = parent_path
    workbook_path = str(workbook_path) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    ###volume###
        #no data
    ###tube###
        #no data
    ###save excel###
    workbook.save(workbook_path)

def step_10(file_name, column, parent_path):
    ###workbook###
    workbook_path = parent_path
    workbook_path = str(workbook_path) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    ###volume###
    #dna#
    #5X KCM
    fiveX_KCM_per_tube = float(ws["E92"].value)
    ws['E92'] = fiveX_KCM_per_tube / column
    fiveX_KCM_total = float(ws["E93"].value)
    ws['E93'] = fiveX_KCM_total / column
    #E. coli competent cell
    ecoli_per_tube = float(ws["E95"].value)
    ws['E95'] = ecoli_per_tube / column
    ecoli_total = float(ws["E96"].value)
    ws['E96'] = ecoli_total / column
    #DW
    dw_dna = float(ws["E110"].value)
    ws['E110'] = dw_dna / column
    #cell#
    #LB medium
    LB_medium = float(ws["E187"].value)
    ws['E187'] = LB_medium / column
    ###tube###
    if 1<= column <= 6:
        ws['B30'] = " "
        ws['B30'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
        ws['B28'] = " "
        ws['B28'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    ###save excel###
    workbook.save(workbook_path)

def step_11_KCM(file_name, column, parent_path):
    workbook_path = parent_path
    workbook_path = str(workbook_path) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    if 1 <= column <= 6:
        ws['B28'] = " "
        ws['B28'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    ###save excel###
    workbook.save(workbook_path)
        
def step_11_ecoli(file_name, column, parent_path):
    workbook_path = parent_path
    workbook_path = str(workbook_path) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    if 1 <= column <= 6:
        ws['B30'] = " "
        ws['B30'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    ###save excel###
    workbook.save(workbook_path)


def step_11(file_name, column, parent_path):
    ###workbook###
    workbook_path = parent_path
    workbook_path = str(workbook_path) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    ###volume###
    #dna#
    #insert dna
    insert_dna_per_tube = float(ws["E92"].value)
    ws['E92'] = insert_dna_per_tube / column
    insert_dna_total = float(ws["E93"].value)
    ws['E93'] = insert_dna_total / column
    #golden gate mixture
    template_dna = float(ws["E96"].value)
    ws['E96'] = insert_dna_total / column
    t4_dna_ligase_buffer = float(ws["E97"].value)
    ws['E97'] = t4_dna_ligase_buffer / column
    neb_golden_gate_enzyme_premix = float(ws["E98"].value)
    ws['E98'] = neb_golden_gate_enzyme_premix / column
    golden_gate_dw = float(ws["E99"].value)
    ws['E99'] = golden_gate_dw / column
    golden_gate_mixture_total = float(ws["E100"].value)
    ws['E100'] = golden_gate_mixture_total / column
    #5X kCM
    fiveX_KCM_per_tube = float(ws["E103"].value)
    ws['E103'] = fiveX_KCM_per_tube / column
    fiveX_KCM_total = float(ws["E104"].value)
    ws['E104'] = fiveX_KCM_total / column
    #E.coli competent cell
    ecoli_per_tube = float(ws["E106"].value)
    ws['E106'] = ecoli_per_tube / column
    ecoli_total = float(ws["E107"].value)
    ws['E107'] = ecoli_total / column
    #D.W.
    transformation_dw = float(ws["E112"].value)
    ws['E112'] = transformation_dw / column
    #cell#
    #LB medium
    LB_medium = float(ws["E187"].value)
    ws['E187'] = LB_medium / column   
    ###save excel###
    workbook.save(workbook_path)
