#created on: 2021 12 25
#Latest Update on: 2022 05 03
#author: github.com/F0RTRE55

import os.path
import xlsxwriter
import openpyxl
from openpyxl import load_workbook
from win32com.client import Dispatch

#create destination file
def create_destination_file(current_directory, file_name):
    directory = str(current_directory)+"\\result\\"+file_name+".xlsx"
    workbook = xlsxwriter.Workbook(directory)
    workbook.close()

#create past order file
def create_past_order_file(current_directory):
    past_order_directory = str(current_directory) + "/past_orders.xlsx"
    if os.path.exists(past_order_directory) == 1:
        return 0
    else:
        #create excel
        past_order_workbook = xlsxwriter.Workbook(past_order_directory)
        past_order_workbook.close()
        #copy format
        format_directory = str(current_directory) + "/format.xlsx"
        excel = Dispatch("Excel.Application") #find program in window that can handle excel
        excel.Visible = False #hide executed excel file
        format_workbook = excel.Workbooks.Open(Filename = format_directory)
        past_order_workbook = excel.Workbooks.Open(Filename = past_order_directory)
        format_sheet = format_workbook.Worksheets("past_order")
        format_sheet.Copy(Before = past_order_workbook.Worksheets(1))
        past_order_workbook.Close(SaveChanges = True)
        excel.Quit()

#copy step format to the destination file
def copy_file(current_directory, file_name, step):
    #copy well format
    format_directory = str(current_directory) + "/format.xlsx"
    destination_directory = str(current_directory)+"\\result\\"+file_name+".xlsx"
    format_directory = str(current_directory) + "/format.xlsx"
    excel = Dispatch("Excel.Application") #find program in window that can handle excel
    excel.Visible = False #hide executed excel file
    format_workbook = excel.Workbooks.Open(Filename = format_directory)
    past_order_workbook = excel.Workbooks.Open(Filename = destination_directory)
    format_sheet = format_workbook.Worksheets("well")
    format_sheet.Copy(Before = past_order_workbook.Worksheets(1))
    past_order_workbook.Close(SaveChanges = True)
    excel.Quit()
    #copy step format
    format_directory = str(current_directory) + "/format.xlsx"
    input_step = "step " + str(step)
    destination_directory = str(current_directory)+"\\result\\"+file_name+".xlsx"
    format_directory = str(current_directory) + "/format.xlsx"
    excel = Dispatch("Excel.Application") #find program in window that can handle excel
    excel.Visible = False #hide executed excel file
    format_workbook = excel.Workbooks.Open(Filename = format_directory)
    past_order_workbook = excel.Workbooks.Open(Filename = destination_directory)
    format_sheet = format_workbook.Worksheets(input_step)
    format_sheet.Copy(Before = past_order_workbook.Worksheets(1))
    past_order_workbook.Close(SaveChanges = True)
    excel.Quit()
    