from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def step1(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
        #no changes applied
    ###save excel###
    workbook.save(workbook_path)

def step2(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
    if 1 <= order_volume/8 <= 4:
        ws['F24'] = " "
        ws['F24'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
        ws['F23'] = " "
        ws['F23'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
        ws['F22'] = " "
        ws['F22'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    elif 5 <= order_volume/8 <= 8:
        ws['F24'] = " "
        ws['F24'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
        ws['F23'] = " "
        ws['F23'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    ###save excel###
    workbook.save(workbook_path)

def step3(current_directory, order_volume, file_name):
    step_3_KCM(current_directory, order_volume, file_name)
    step_3_ecoli(current_directory, order_volume, file_name)
    step_3_ligation(current_directory, order_volume, file_name)
def step_3_KCM(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
    if 1 <= order_volume/8 <= 8:
        ws['B28'] = " "
        ws['B28'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    ###save excel###
    workbook.save(workbook_path)
def step_3_ecoli(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
    if 1 <= order_volume/8 <= 6:
        ws['B30'] = " "
        ws['B30'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    ###save excel###
    workbook.save(workbook_path)
def step_3_ligation(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
    if 1 <= order_volume/8 <= 4:
        ws['F28'] = " "
        ws['F28'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
        ws['F27'] = " "
        ws['F27'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
        ws['F26'] = " "
        ws['F26'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    elif 5 <= order_volume/8 <= 8:
        ws['F28'] = " "
        ws['F28'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
        ws['F27'] = " "
        ws['F27'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    ###save excel###
    workbook.save(workbook_path)

def step4(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
        #no changes applied
    ###save excel###
    workbook.save(workbook_path)

def step5(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
        #no changes applied
    ###save excel###
    workbook.save(workbook_path)

def step6(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
        #no changes applied
    ###save excel###
    workbook.save(workbook_path)

def step7(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
        #no changes applied
    ###save excel###
    workbook.save(workbook_path)

def step8(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
    if 1 <= order_volume/8 <= 6:
        ws['B22'] = " "
        ws['B22'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    ###save excel###
    workbook.save(workbook_path)

def step9(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
        #no changes applied
    ###save excel###
    workbook.save(workbook_path)

def step10(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\\\result\\\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
    if 1<= order_volume/8 <= 6:
        ws['B30'] = " "
        ws['B30'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
        ws['B28'] = " "
        ws['B28'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    ###save excel###
    workbook.save(workbook_path)

def step11(current_directory, order_volume, file_name):
    step_11_KCM(current_directory, order_volume, file_name)
    step_11_ecoli(current_directory, order_volume, file_name)
def step_11_KCM(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
    if 1 <= order_volume/8 <= 6:
        ws['B28'] = " "
        ws['B28'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    ###save excel###
    workbook.save(workbook_path)
def step_11_ecoli(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\\\result\\\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
    if 1 <= order_volume/8 <= 6:
        ws['B30'] = " "
        ws['B30'].fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type = "solid")
    ###save excel###
    workbook.save(workbook_path)
