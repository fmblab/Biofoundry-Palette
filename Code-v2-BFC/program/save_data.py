#created on: 2021 12 25
#Latest Update on: 2022 05 03
#author: github.com/F0RTRE55

import openpyxl
from openpyxl import load_workbook

def find_empty_line(current_directory):
    database_path = str(current_directory) + "/past_orders.xlsx"
    workbook = openpyxl.load_workbook(database_path)
    sheetnames = workbook.sheetnames
    n = 0
    while sheetnames[n] != "past_order":
        n += 1
    ws = workbook.worksheets[n]
    for cell in ws ["C"]:
        if cell.value is None:
            return cell.row
    else:
        return cell.row + 1

def save_protocol(current_directory, row, operator, client, order_volume, orderID, step, order_date, deadline, file_name):
    database_path = str(current_directory) + "/past_orders.xlsx"
    #finding sheet
    workbook = openpyxl.load_workbook(database_path)
    sheetnames = workbook.sheetnames
    n = 0
    while sheetnames[n] != "past_order":
        n += 1
    workbook.close()
    #workbook
    workbook = load_workbook(database_path)
    ws = workbook.active
    ws = workbook.worksheets[n]
    #operator name
    row = str(row)
    ws['A'+row] = operator
    #client name
    ws['B'+row] = client
    #order date
    ws['C'+row] = order_date
    #due date
    ws['D'+row] = deadline
    #alt. gene id
    ws['E'+row] = orderID
    #order number
    ws['F'+row] = order_volume
    #step
    ws['G'+row] = step
    #file name
    ws['H'+row] = file_name
    #save excel
    workbook.save(database_path)
