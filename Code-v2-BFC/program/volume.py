#created on: 2021 12 25
#Latest Update on: 2022 05 03
#author: github.com/F0RTRE55

from re import S
from openpyxl import load_workbook

def step1(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\\\result\\\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
    #Elution buffer
    elution_per_tube = float(ws["E92"].value)
    ws['E92'] = elution_per_tube * order_volume
    elution_total = float(ws["E93"].value)
    ws['E93'] = elution_total * order_volume / 8
    #Buffer PM
    buffer_pm = float(ws["E105"].value)
    ws['E105'] = buffer_pm * order_volume / 8
    #Buffer PE
    buffer_pe_per_tube = float(ws["E107"].value)
    ws['E107'] = buffer_pe_per_tube * order_volume / 8
    buffer_pe_total = float(ws["E108"].value)
    ws['E108'] = buffer_pe_total * order_volume / 8
    #GXL PCR pre-mix
    template_dna = float(ws["E111"].value)
    ws['E111'] = template_dna * order_volume / 8
    reverse_primer =  float(ws["E112"].value)
    ws['E112'] = reverse_primer * order_volume / 8
    dNTP =  float(ws["E113"].value)
    ws['E113'] = dNTP * order_volume / 8
    fiveX_GXL_buffer =  float(ws["E114"].value)
    ws['E114'] = fiveX_GXL_buffer * order_volume / 8
    GXL_polymerase =  float(ws["E115"].value)
    ws['E115'] = GXL_polymerase * order_volume / 8
    DW =  float(ws["E116"].value)
    ws['E116'] = DW * order_volume / 8
    GXL_total =  float(ws["E117"].value)
    ws['E117'] = GXL_total * order_volume / 8
    ###save excel###
    workbook.save(workbook_path)

def step2(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\\\result\\\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
    elution_buffer_per_tube =  float(ws["E92"].value)
    ws['E92'] = elution_buffer_per_tube * order_volume / 8
    elution_buffer_total =  float(ws["E93"].value)
    ws['E93'] = elution_buffer_total * order_volume / 8
    #enzyme cut
    tenX_green_fd_buffer =  float(ws["E96"].value)
    ws['E96'] = tenX_green_fd_buffer * order_volume / 8
    spel =  float(ws["E97"].value)
    ws['E97'] = spel * order_volume / 8
    distilled_water =  float(ws["E98"].value)
    ws['E98'] = distilled_water * order_volume / 8
    enzyme_per_tube =  float(ws["E99"].value)
    if 1 <= order_volume/8 <= 4:
        ws['E99'] = enzyme_per_tube * order_volume / 8
    elif 5 <= order_volume/8 <= 8:
        ws['E99'] = enzyme_per_tube * order_volume / 8 / 2
    elif 9 <= order_volume/8 <= 12:
        ws['E99'] = enzyme_per_tube * order_volume / 8 / 4
    enzyme_total =  float(ws["E100"].value)
    ws['E100'] = enzyme_total * order_volume / 8
    #buffer PM
    buffer_pm =  float(ws["E105"].value)
    ws['E105'] = buffer_pm * order_volume / 8
    #buffer PE
    buffer_pe_per_trough =  float(ws["E107"].value)
    ws['E107'] = buffer_pe_per_trough * order_volume / 8
    buffer_pe_total =  float(ws["E108"].value)
    ws['E108'] = buffer_pe_total * order_volume / 8
    ###save excel###
    workbook.save(workbook_path)

def step3(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\\\result\\\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
    #DNA#
    #5X KCM
    fiveX_KCM_per_tube =  float(ws["E92"].value)
    if 1 <= order_volume/8 <= 8:
        ws['E92'] = fiveX_KCM_per_tube * order_volume / 8
    elif 9 <= order_volume/8 <= 12:
        ws['E92'] = fiveX_KCM_per_tube * order_volume / 8 / 2
    fiveX_KCM_total =  float(ws["E93"].value)
    ws['E93'] = fiveX_KCM_total * order_volume / 8
    #E.coli competent cell
    ecoli_per_tube =  float(ws["E95"].value)
    if 1 <= order_volume/8 <= 6:
        ws['E95'] = ecoli_per_tube * order_volume / 8
    elif 1 <= order_volume/8 <= 7:
        ws['E95'] = ecoli_per_tube * order_volume / 8 / 2
    ecoli_total =  float(ws["E96"].value)
    ws['E96'] = ecoli_total * order_volume / 8
    #ligation
    tenX_ligation =  float(ws["E99"].value)
    ws['E99'] = tenX_ligation * order_volume / 8
    T4_ligase =  float(ws["E100"].value)
    ws['E100'] = T4_ligase * order_volume / 8
    DW_tube =  float(ws["E101"].value)
    ws['E101'] = DW_tube * order_volume / 8
    ligation_pre_mix_per_tube =  float(ws["E102"].value)
    if 1 <= order_volume/8 <= 4:
        ws['E102'] = ligation_pre_mix_per_tube * order_volume / 8
    elif 5 <= order_volume/8 <= 8:
        ws['E102'] = ligation_pre_mix_per_tube * order_volume / 8 / 2
    elif 9 <= order_volume/8 <= 12:
        ws['E102'] = ligation_pre_mix_per_tube * order_volume / 8 / 4
    ligation_pre_mix_total =  float(ws["E103"].value)
    ws['E103'] = ligation_pre_mix_total * order_volume / 8
    #d.w.
    DW_trough =  float(ws["E110"].value)
    ws['E110'] = DW_trough * order_volume / 8
    #CELL#
    #LB medium
    LB_medium =  float(ws["E187"].value)
    ws['E187'] = LB_medium * order_volume / 8
    ###save excel###
    workbook.save(workbook_path)
    
def step4(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
    #dw
    dw = float(ws["E58"].value)
    ws['E58'] = dw * order_volume / 8
    #lb_medium
    lb_medium = float(ws["E61"].value)
    ws['E61'] = lb_medium * order_volume / 8
    ###save excel###
    workbook.save(workbook_path)

def step5(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\\\result\\\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
    #DW
    dw_per_tube = float(ws["E91"].value)
    ws['E91'] = dw_per_tube * order_volume / 8
    dw_total =  float(ws["E92"].value)
    ws['E92'] = dw_total * order_volume / 8
    #buffer p1
    bufferp1 =  float(ws["E104"].value)
    ws['E104'] = bufferp1 * order_volume / 8
    #buffer p2
    bufferp2 =  float(ws["E106"].value)
    ws['E106'] = bufferp2 * order_volume / 8
    #buffer N3
    buffern3 =  float(ws["E108"].value)
    ws['E108'] = buffern3 * order_volume / 8
    #buffer pb
    bufferpb =  float(ws["E110"].value)
    ws['E110'] = bufferpb * order_volume / 8
    #buffer pe
    bufferpe_per_trough =  float(ws["E112"].value)
    ws['E112'] = bufferpe_per_trough * order_volume / 8
    bufferpe_total =  float(ws["E113"].value)
    ws['E113'] = bufferpe_total * order_volume / 8
    ###save excel###
    workbook.save(workbook_path)

def step6(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\\\result\\\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
    #DNA#
        #no volume change
    #cell#
    BHISG = float(ws["E186"].value)
    ws['E186'] = BHISG * order_volume / 8
    ###save excel###
    workbook.save(workbook_path)

def step7(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\\\result\\\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
    #BHIS
    BHIS_medium = float(ws["E64"].value)
    ws['E64'] = BHIS_medium * order_volume / 8
    #dw
    dw = float(ws["E66"].value)
    ws['E66'] = dw * order_volume / 8
    ###save excel###
    workbook.save(workbook_path)

def step8(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\\\result\\\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
    #2X sapphire
    twoX_sapphire_per_tube = float(ws["E91"].value)
    if 1 <= order_volume/8 <= 6:
        ws['E91'] = twoX_sapphire_per_tube * order_volume / 8
    elif 7 <= order_volume/8 <= 12:
        ws['E91'] = twoX_sapphire_per_tube * order_volume / 8 / 2
    twoX_sapphire_total = float(ws["E92"].value)
    ws['E92'] = twoX_sapphire_total * order_volume / 8
    #buffer EB
    buffer_eb_per_tube = float(ws["E94"].value)
    if 1 <= order_volume/8 <= 6:
        ws['E94'] = buffer_eb_per_tube * order_volume / 8
    elif 7 <= order_volume/8 <= 12:
        ws['E91'] = twoX_sapphire_per_tube * order_volume / 8 / 2
    buffer_eb_total = float(ws["E95"].value)
    ws['E95'] = buffer_eb_total * order_volume / 8
    #buffer PM
    buffer_pm = float(ws["E104"].value)
    ws['E104'] = buffer_pm * order_volume / 8
    #buffer PE
    buffer_pe_per_tube = float(ws["E106"].value)
    ws['E106'] = buffer_pe_per_tube * order_volume / 8
    buffer_pe_total = float(ws["E107"].value)
    ws['E107'] = buffer_pe_total * order_volume / 8
    ###save excel###
    workbook.save(workbook_path)

def step9(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\\\result\\\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
        #no changes applied
    ###save excel###
    workbook.save(workbook_path)

def step10(current_directory, order_volume, file_name):###workbook###
    workbook_path = str(current_directory) + "\\\\result\\\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
    #dna#
    #5X KCM
    fiveX_KCM_per_tube = float(ws["E92"].value)
    ws['E92'] = fiveX_KCM_per_tube * order_volume / 8
    fiveX_KCM_total = float(ws["E93"].value)
    ws['E93'] = fiveX_KCM_total * order_volume / 8
    #E. coli competent cell
    ecoli_per_tube = float(ws["E95"].value)
    ws['E95'] = ecoli_per_tube * order_volume / 8
    ecoli_total = float(ws["E96"].value)
    ws['E96'] = ecoli_total * order_volume / 8
    #DW
    dw_dna = float(ws["E110"].value)
    ws['E110'] = dw_dna * order_volume / 8
    #cell#
    #LB medium
    LB_medium = float(ws["E187"].value)
    ws['E187'] = LB_medium * order_volume / 8
    ###save excel###
    workbook.save(workbook_path)

def step11(current_directory, order_volume, file_name):
    ###workbook###
    workbook_path = str(current_directory) + "\\\\result\\\\" + file_name +".xlsx"
    workbook = load_workbook(workbook_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #rules
    #dna#
    #insert dna
    insert_dna_per_tube = float(ws["E92"].value)
    ws['E92'] = insert_dna_per_tube * order_volume / 8
    insert_dna_total = float(ws["E93"].value)
    ws['E93'] = insert_dna_total * order_volume / 8
    #golden gate mixture
    template_dna = float(ws["E96"].value)
    ws['E96'] = template_dna * order_volume / 8
    t4_dna_ligase_buffer = float(ws["E97"].value)
    ws['E97'] = t4_dna_ligase_buffer * order_volume / 8
    neb_golden_gate_enzyme_premix = float(ws["E98"].value)
    ws['E98'] = neb_golden_gate_enzyme_premix * order_volume / 8
    golden_gate_dw = float(ws["E99"].value)
    ws['E99'] = golden_gate_dw * order_volume / 8
    golden_gate_mixture_total = float(ws["E100"].value)
    ws['E100'] = golden_gate_mixture_total * order_volume / 8
    #5X kCM
    fiveX_KCM_per_tube = float(ws["E103"].value)
    if 1 <= order_volume/8 <= 6:
        ws['E103'] = fiveX_KCM_per_tube * order_volume / 8
    elif 7 <= order_volume/8 <= 12:
        ws['E103'] = fiveX_KCM_per_tube * order_volume / 8 / 2
    fiveX_KCM_total = float(ws["E104"].value)
    ws['E104'] = fiveX_KCM_total * order_volume / 8
    #E.coli competent cell
    ecoli_per_tube = float(ws["E106"].value)
    if 1 <= order_volume/8 <= 6:
        ws['E106'] = ecoli_per_tube * order_volume / 8
    elif 7 <= order_volume/8 <= 12:
        ws['E106'] = ecoli_per_tube * order_volume / 8 / 2
    ecoli_total = float(ws["E107"].value)
    ws['E107'] = ecoli_total * order_volume / 8
    #D.W.
    transformation_dw = float(ws["E112"].value)
    ws['E112'] = transformation_dw * order_volume / 8
    #cell#
    #LB medium
    LB_medium = float(ws["E187"].value)
    ws['E187'] = LB_medium * order_volume / 8   
    ###save excel###
    workbook.save(workbook_path)