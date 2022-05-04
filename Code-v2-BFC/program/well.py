#created on: 2021 12 25
#Latest Update on: 2022 05 03
#author: github.com/F0RTRE55

from openpyxl import load_workbook
from program import adder

#new well plate
def new_well(current_directory, order_volume, wellID, well_position, file_name):
    #calculatation
    column_whole = int(order_volume/8.0)
    column_des = float(str(order_volume/8-int(order_volume/8))[1:])
    column_des = int(column_des * 8)
    position = int(well_position)
    repeating_wells(file_name, position, column_whole, column_des, current_directory)
    adder.save_well_data(column_whole, column_des)

def repeating_wells (name, position, column_whole, column_des, parent_path):
    result_excel_path = parent_path
    result_excel_path = str(result_excel_path) + "\\result\\" + name + ".xlsx"
    workbook = load_workbook(result_excel_path)
    ws = workbook.active
    ws = workbook.worksheets[1]
    for i in range (0, column_whole):
        if position == 1:
            cell_pos = "C"
        elif position == 2:
            cell_pos = "D"
        elif position == 3:
            cell_pos = "E"
        elif position == 4:
            cell_pos = "F"
        elif position == 5:
            cell_pos = "G"
        elif position == 6:
            cell_pos = "H"
        elif position == 7:
            cell_pos = "I"
        elif position == 8:
            cell_pos = "J"
        elif position == 9:
            cell_pos = "K"
        elif position == 10:
            cell_pos = "L"
        elif position == 11:
            cell_pos = "M"
        elif position == 12:
            cell_pos = "N"
        elif position == 13:
            cell_pos = "O"
        elif position == 14:
            cell_pos = "P"
        elif position == 15:
            cell_pos = "Q"
        elif position == 16:
            cell_pos = "R"
        elif position == 17:
            cell_pos = "S"
        elif position == 18:
            cell_pos = "T"
        elif position == 19:
            cell_pos = "U"
        elif position == 20:
            cell_pos = "V"
        elif position == 21:
            cell_pos = "W"
        elif position == 22:
            cell_pos = "X"
        elif position == 23:
            cell_pos = "Y"
        elif position == 24:
            cell_pos = "Z"
        else:
            print("something is wrong")
        cell_pos_row = 6
        for j in range (1, 9):
            cell_comp = cell_pos + str(cell_pos_row)
            ws[cell_comp] = "X"
            cell_pos_row = cell_pos_row + 1
        position = position + 1
    #save excel
    workbook.save(result_excel_path)
    single_wells(name, position, column_des, parent_path)

def single_wells(name, position, column_des, parent_path):
    result_excel_path = parent_path
    result_excel_path = str(result_excel_path) + "\\result\\" + name + ".xlsx"
    workbook = load_workbook(result_excel_path)
    ws = workbook.active
    ws = workbook.worksheets[1]
    cell_pos_row = 6
    for i in range(0, column_des):
        cell_pos_2 = "something"
        if position == 1:
            cell_pos_2 = "C"
        elif position == 2:
            cell_pos_2 = "D"
        elif position == 3:
            cell_pos_2 = "E"
        elif position == 4:
            cell_pos_2 = "F"
        elif position == 5:
            cell_pos_2 = "G"
        elif position == 6:
            cell_pos_2 = "H"
        elif position == 7:
            cell_pos_2 = "I"
        elif position == 8:
            cell_pos_2 = "J"
        elif position == 9:
            cell_pos_2 = "K"
        elif position == 10:
            cell_pos_2 = "L"
        elif position == 11:
            cell_pos_2 = "M"
        elif position == 12:
            cell_pos_2 = "N"
        elif position == 13:
            cell_pos_2 = "O"
        elif position == 14:
            cell_pos_2 = "P"
        elif position == 15:
            cell_pos_2 = "Q"
        elif position == 16:
            cell_pos_2 = "R"
        elif position == 17:
            cell_pos_2 = "S"
        elif position == 18:
            cell_pos_2 = "T"
        elif position == 19:
            cell_pos_2 = "U"
        elif position == 20:
            cell_pos_2 = "V"
        elif position == 21:
            cell_pos_2 = "W"
        elif position == 22:
            cell_pos_2 = "X"
        elif position == 23:
            cell_pos_2 = "Y"
        elif position == 24:
            cell_pos_2 = "Z"
        else:
            print("something is wrong")
        cell_comp = cell_pos_2 + str(cell_pos_row)
        cell_pos_row = cell_pos_row + 1        
        ws[cell_comp] = "X"
        workbook.save(result_excel_path)
