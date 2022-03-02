import os, sys, openpyxl, datetime, xlsxwriter
from openpyxl import load_workbook
from win32com.client import Dispatch
from tkinter import messagebox
from pathlib import Path
    
def main (operator, client, order_number, order_name, step, order_date, due_date, file_name, parent_path):
    ask_question()
    database_path = check_step(step, parent_path)
    file_exist(database_path, parent_path)
    due_year = convert_to_year(due_date)
    check_sheet(database_path, due_year, parent_path)
    row = check_empty(database_path, due_year)
    save_protocol(operator, client, order_number, order_name, step, order_date, due_date, file_name, database_path, row, due_year)
    done_message()

def ask_question():
    reply = messagebox.askquestion("BFC V.4","Do you want to save order information?")
    if reply == "yes":
        return 1
    else:
        exit()

def check_step(step, parent_path):
    parent_directory = parent_path
    if step == "11":
        return str(parent_directory) + "\past_orders_step_11.xlsx"
    else:
        return str(parent_directory) + "\past_orders_others.xlsx"

def file_exist(database_path, parent_path):
    existance = os.path.exists(database_path)
    if existance == 1:
        return 0
    else:
        create_excel(database_path, parent_path)

def create_excel(database_path, parent_path):
    #create excel
    workbook = xlsxwriter.Workbook(database_path)
    workbook.close()
    #get year
    now = datetime.datetime.now()
    #copy sheet
    format_path = str(parent_path) + "\\format_excels\past_orders_template.xlsx"
    xl = Dispatch("Excel.Application")
    xl.Visible = False
    format_workbook = xl.Workbooks.Open(Filename = format_path)
    dest_workbook = xl.Workbooks.Open(Filename = database_path)
    format_sheet = format_workbook.Worksheets('Sheet1')
    format_sheet.Copy(Before = dest_workbook.Worksheets(1))
    dest_workbook.Close(SaveChanges = True)
    xl.Quit()
    #change sheet name
    ss = openpyxl.load_workbook(database_path)
    ss_sheet = ss['Sheet1 (2)']
    ss_sheet.title = str(now.year)
    ss.save(database_path)

def convert_to_year(due_date):
    due_date = str(due_date)
    due_date = due_date[-2:]
    due_year = "20" + str(due_date)
    return due_year

def check_sheet(database_path, due_year, parent_path):
    workbook = openpyxl.load_workbook(database_path)
    if due_year in workbook.sheetnames:
        return 0
    else:
        create_sheet_year(database_path, due_year, parent_path)
        return 0

def create_sheet_year(database_path, due_year, parent_path):
    #get year
    now = datetime.datetime.now()
    #copy sheet
    format_path = str(parent_path) + "\\format_excels\past_orders_template.xlsx"
    xl = Dispatch("Excel.Application")
    xl.Visible = False
    format_workbook = xl.Workbooks.Open(Filename = format_path)
    dest_workbook = xl.Workbooks.Open(Filename = database_path)
    format_sheet = format_workbook.Worksheets('Sheet1')
    format_sheet.Copy(Before = dest_workbook.Worksheets(1))
    dest_workbook.Close(SaveChanges = True)
    xl.Quit()
    #change sheet name
    ss = openpyxl.load_workbook(database_path)
    ss_sheet = ss['Sheet1 (2)']
    ss_sheet.title = str(due_year)
    ss.save(database_path)

def check_empty(database_path, due_year):
    workbook = openpyxl.load_workbook(database_path)
    sheetnames = workbook.sheetnames
    n = 0
    while sheetnames[n] != due_year:
        n += 1
    ws = workbook.worksheets[n]
    for cell in ws ["C"]:
        if cell.value is None:
            return cell.row
    else:
        return cell.row + 1

def save_protocol(operator, client, order_number, order_name, step, order_date, due_date, file_name, database_path, row, due_year):
    #finding sheet
    workbook = openpyxl.load_workbook(database_path)
    sheetnames = workbook.sheetnames
    n = 0
    while sheetnames[n] != due_year:
        n += 1
    workbook.close()
    #workbook
    workbook = load_workbook(database_path)
    ws = workbook.active
    ws = workbook.worksheets[n]
    #operator name
    row = str(row)
    ws['A'+row] = operator
    #client name
    ws['B'+row] = client
    #order date
    ws['C'+row] = order_date
    #due date
    ws['D'+row] = due_date
    #alt. gene id
    ws['E'+row] = order_name
    #order number
    ws['F'+row] = order_number
    #step
    ws['G'+row] = step
    #file name
    ws['H'+row] = file_name
    #save excel
    workbook.save(database_path)

def done_message():
    messagebox.askquestion("BFC V.4","Data Saved")
