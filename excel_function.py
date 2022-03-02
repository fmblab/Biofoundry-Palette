import xlsxwriter, os, sys
from pathlib import Path
from openpyxl import load_workbook
from win32com.client import Dispatch
from programming import auto_volume

def create_excel(name, parent_path):
    result_path = parent_path
    result_path = str(result_path) + "\\result\\" + name + ".xlsx"
    workbook = xlsxwriter.Workbook(result_path)
    workbook.close()

def copy_step_format(name, step, parent_path):
    format_path = parent_path
    format_path = str(format_path) + "\\format_excels\\" + "formatting_full_final.xlsx"
    destination_path = parent_path
    destination_path = str(destination_path) + "\\result\\" + name + ".xlsx"
    xl = Dispatch("Excel.Application")
    xl.Visible = False
    format_workbook = xl.Workbooks.Open(format_path) #open formatting file
    destination_workbook = xl.Workbooks.Open(destination_path) #open destination file
    sheet = "step " + str(step)
    format_step_worksheet = format_workbook.Worksheets(sheet) #search step
    format_well_worksheet = format_workbook.Worksheets("well") #search well
    format_step_worksheet.Copy(Before = destination_workbook.Worksheets(1)) #copy step format
    format_well_worksheet.Copy(Before = destination_workbook.Worksheets(2)) #copy well format
    destination_workbook.Close(SaveChanges = True)
    xl.Quit()

def basic_info_to_copied_format(operator, client, order_number, order_name, order_date, due_date, file_name, parent_path):
    #open excel
    result_excel_path = parent_path
    result_excel_path = str(result_excel_path) + "\\result\\" + file_name +".xlsx"
    workbook = load_workbook(result_excel_path)
    ws = workbook.active
    ws = workbook.worksheets[0]
    #operator name
    ws['C3'] = operator
    #client name
    ws['H3'] = client
    #order number
    ws['I8'] = order_number
    #order name
    ws['D7'] = order_name
    #order date
    ws['C4'] = order_date
    #due date
    ws['H4'] = due_date
    #save excel
    workbook.save(result_excel_path)

def volume_calculation(file_name, step, column_height, parent_path):
    auto_volume.main(file_name, step, column_height, parent_path)
